from flask import Flask, render_template, request, jsonify, redirect, url_for, session, send_from_directory, make_response
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from datetime import datetime, timezone, timedelta
import os
import configparser
import uuid
import requests
import json
import re
import threading
from threading import Lock
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key-change-in-production'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///homework_system.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {'pool_pre_ping': True, 'pool_recycle': 300}

# 读取配置文件
config = configparser.ConfigParser()
config.read('homework.ini', encoding='utf-8')

# 图片上传配置
ENABLE_IMAGE_UPLOAD = config.getboolean('settings', 'enable_image_upload', fallback=False)
MAX_IMAGES_PER_HOMEWORK = config.getint('settings', 'max_images_per_homework', fallback=5)
ALLOWED_EXTENSIONS = set(config.get('settings', 'allowed_image_formats', fallback='jpg,jpeg,png,gif').split(','))
MAX_IMAGE_SIZE_MB = config.getint('settings', 'max_image_size_mb', fallback=10)

# AI复核配置
ENABLE_AI_REVIEW = config.getboolean('ai_review', 'enable_ai_review', fallback=False)
AI_API_URL = config.get('ai_review', 'ai_api_url', fallback='https://ack-ai.qinyining.cn/pg/chat/completions')
AI_MODEL = config.get('ai_review', 'ai_model', fallback='gpt-4.1-mini')
HOMEWORK_BASE_URL = config.get('ai_review', 'homework_base_url', fallback='https://tmptest.qinyining.cn')
AI_REVIEW_ACTION = config.get('ai_review', 'ai_review_action', fallback='mark_abnormal')
AI_REVIEW_MAX_RETRIES = config.getint('ai_review', 'ai_review_max_retries', fallback=3)

# AI API认证信息
AI_LOGIN_URL = 'https://qin.qinyining.cn/api/user/login?turnstile='
AI_USERNAME = 'private'
AI_PASSWORD = 'password'

# 全局cookie存储
ai_session_cookie = None
ai_cookie_lock = Lock()

# 图片上传目录
UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = MAX_IMAGE_SIZE_MB * 1024 * 1024

db = SQLAlchemy(app)

# 设置中国时区 UTC+8
CHINA_TZ = timezone(timedelta(hours=8))

def get_china_time():
    """获取中国时间（UTC+8）"""
    return datetime.now(CHINA_TZ)

def allowed_file(filename):
    """检查文件扩展名是否允许"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# 数据库模型
class Admin(db.Model):
    """管理员表"""
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password = db.Column(db.String(200), nullable=False)
    created_at = db.Column(db.DateTime, default=get_china_time)

class Teacher(db.Model):
    """教师表"""
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password = db.Column(db.String(200), nullable=False)
    subject = db.Column(db.String(50), nullable=False)  # 学科
    enable_ai_review = db.Column(db.Boolean, default=True)  # 是否启用AI复审
    created_at = db.Column(db.DateTime, default=get_china_time)

class Student(db.Model):
    """学生表"""
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(80), nullable=False)
    student_id = db.Column(db.String(50), unique=True, nullable=False)
    created_at = db.Column(db.DateTime, default=get_china_time)

class Homework(db.Model):
    """作业布置表"""
    id = db.Column(db.Integer, primary_key=True)
    subject = db.Column(db.String(50), nullable=False)  # 学科
    teacher_id = db.Column(db.Integer, db.ForeignKey('teacher.id'), nullable=False)
    title = db.Column(db.String(200), nullable=False)  # 作业标题
    ai_prompt = db.Column(db.Text)  # 自定义AI检测prompt
    max_images = db.Column(db.Integer, default=5)  # 允许上传的最大图片数量
    created_at = db.Column(db.DateTime, default=get_china_time)
    teacher = db.relationship('Teacher', backref='homeworks')

class HomeworkSubmission(db.Model):
    """作业提交记录表"""
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.Integer, db.ForeignKey('student.id'), nullable=False)
    homework_id = db.Column(db.Integer, db.ForeignKey('homework.id'), nullable=False)
    submitted_at = db.Column(db.DateTime, default=get_china_time)
    ai_review_status = db.Column(db.String(20), default='pending')  # pending, reviewing, approved, rejected, error
    ai_review_result = db.Column(db.Text)  # AI审核的详细结果
    ai_reviewed_at = db.Column(db.DateTime)  # AI审核时间
    student = db.relationship('Student', backref='submissions')
    homework = db.relationship('Homework', backref='submissions')

class HomeworkImage(db.Model):
    """作业图片表"""
    id = db.Column(db.Integer, primary_key=True)
    submission_id = db.Column(db.Integer, db.ForeignKey('homework_submission.id'), nullable=False)
    filename = db.Column(db.String(200), nullable=False)  # 存储的文件名
    original_filename = db.Column(db.String(200), nullable=False)  # 原始文件名
    uploaded_at = db.Column(db.DateTime, default=get_china_time)
    submission = db.relationship('HomeworkSubmission', backref='images')

# 初始化数据库
with app.app_context():
    db.create_all()
    # 创建默认管理员账户 (admin/admin123)
    admin = Admin.query.filter_by(username='admin').first()
    if not admin:
        admin = Admin(username='admin', password=generate_password_hash('admin123'))
        db.session.add(admin)
        db.session.commit()
        print("[系统] 已创建默认管理员账户: admin/admin123")

# 定时任务：每天00:00清空学生端前一天的作业显示
def clear_previous_day_homework_for_students():
    """清空学生端前一天的作业（仅影响学生端显示，教师端仍可查看）"""
    with app.app_context():
        try:
            now = get_china_time()
            today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
            
            # 获取今天之前布置的所有作业
            old_homeworks = Homework.query.filter(Homework.created_at < today_start).all()
            
            if old_homeworks:
                print(f"[定时任务] {now.strftime('%Y-%m-%d %H:%M:%S')} - 清理前一天作业，共 {len(old_homeworks)} 个作业")
            else:
                print(f"[定时任务] {now.strftime('%Y-%m-%d %H:%M:%S')} - 无需清理")
                
        except Exception as e:
            print(f"[定时任务] 清理作业失败: {str(e)}")
            import traceback
            traceback.print_exc()

# 定时任务：清理无图片的提交记录和超时的判定中状态
def cleanup_invalid_submissions():
    """清理无图片的提交记录，并将超时的判定中状态转为error"""
    with app.app_context():
        try:
            now = get_china_time()
            
            # 1. 删除没有图片的提交记录
            all_submissions = HomeworkSubmission.query.all()
            deleted_count = 0
            for submission in all_submissions:
                image_count = HomeworkImage.query.filter_by(submission_id=submission.id).count()
                if image_count == 0:
                    db.session.delete(submission)
                    deleted_count += 1
            
            if deleted_count > 0:
                db.session.commit()
                print(f"[定时任务] {now.strftime('%Y-%m-%d %H:%M:%S')} - 删除无图片提交记录: {deleted_count} 条")
            
            # 2. 将判定中超过5分钟的记录转为error状态
            timeout_threshold = now - timedelta(minutes=5)
            timeout_submissions = HomeworkSubmission.query.filter(
                HomeworkSubmission.ai_review_status == 'reviewing',
                HomeworkSubmission.submitted_at < timeout_threshold
            ).all()
            
            timeout_count = 0
            for submission in timeout_submissions:
                submission.ai_review_status = 'error'
                submission.ai_review_result = 'AI判定超时（超过5分钟）'
                submission.ai_reviewed_at = now
                timeout_count += 1
            
            if timeout_count > 0:
                db.session.commit()
                print(f"[定时任务] {now.strftime('%Y-%m-%d %H:%M:%S')} - 判定超时转error: {timeout_count} 条")
                
        except Exception as e:
            db.session.rollback()
            print(f"[定时任务] 清理无效提交失败: {str(e)}")
            import traceback
            traceback.print_exc()

# 初始化定时任务调度器
scheduler = BackgroundScheduler(timezone='Asia/Shanghai')

# 每天北京时间00:00执行清理任务
scheduler.add_job(
    func=clear_previous_day_homework_for_students,
    trigger=CronTrigger(hour=0, minute=0, timezone='Asia/Shanghai'),
    id='clear_homework_daily',
    name='清空学生端前一天作业',
    replace_existing=True
)

# 每5分钟执行一次清理无效提交和超时判定的任务
scheduler.add_job(
    func=cleanup_invalid_submissions,
    trigger=CronTrigger(minute='*/5', timezone='Asia/Shanghai'),
    id='cleanup_invalid_submissions',
    name='清理无效提交和超时判定',
    replace_existing=True
)

scheduler.start()
print("[系统] 定时任务调度器已启动")
print("[系统] - 每天00:00清空学生端前一天作业")
print("[系统] - 每5分钟清理无图片提交记录和超时判定")

# ==================== 配置接口 ====================
@app.route('/api/config')
def get_config():
    """获取系统配置"""
    return jsonify({
        'enable_image_upload': ENABLE_IMAGE_UPLOAD,
        'max_images_per_homework': MAX_IMAGES_PER_HOMEWORK,
        'allowed_image_formats': list(ALLOWED_EXTENSIONS),
        'max_image_size_mb': MAX_IMAGE_SIZE_MB,
        'enable_ai_review': ENABLE_AI_REVIEW,
        'ai_review_action': AI_REVIEW_ACTION
    })

# ==================== 学生端路由 ====================
@app.route('/')
def index():
    """学生端首页"""
    return render_template('student.html')

@app.route('/about')
def about():
    """关于页面"""
    return render_template('about.html')

@app.route('/api/students')
def get_students():
    """获取所有学生列表及作业提交状态"""
    students = Student.query.all()
    # 获取所有已布置的作业（仅限当天）
    now = get_china_time()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    today_end = now.replace(hour=23, minute=59, second=59, microsecond=999999)
    
    homeworks = Homework.query.filter(
        Homework.created_at >= today_start,
        Homework.created_at <= today_end
    ).all()
    
    student_list = []
    for student in students:
        # 获取该学生所有学科的提交状态
        homework_status = {}
        for hw in homeworks:
            submission = HomeworkSubmission.query.filter_by(
                student_id=student.id,
                homework_id=hw.id
            ).first()
            
            if hw.subject not in homework_status:
                homework_status[hw.subject] = []
            
            # 获取图片数量和AI审核状态
            image_count = 0
            ai_review_status = None
            ai_review_result = None
            has_images = False
            if submission:
                image_count = HomeworkImage.query.filter_by(submission_id=submission.id).count()
                has_images = image_count > 0
                ai_review_status = submission.ai_review_status
                ai_review_result = submission.ai_review_result
            
            homework_status[hw.subject].append({
                'homework_id': hw.id,
                'title': hw.title,
                'submitted': submission is not None and has_images,  # 只有提交且有图片才算已提交
                'submitted_at': submission.submitted_at.strftime('%Y-%m-%d %H:%M:%S') if submission and has_images else None,
                'submission_id': submission.id if submission else None,
                'image_count': image_count,
                'ai_review_status': ai_review_status,
                'ai_review_result': ai_review_result
            })
        
        student_list.append({
            'id': student.id,
            'name': student.name,
            'student_id': student.student_id,
            'homework_status': homework_status
        })
    
    return jsonify(student_list)

@app.route('/api/create-submission', methods=['POST'])
def create_submission():
    """创建作业提交记录（用于拍照前）"""
    data = request.get_json()
    student_id = data.get('student_id')
    homework_id = data.get('homework_id')
    
    if not student_id:
        return jsonify({'success': False, 'message': '学生ID不能为空'}), 400
    
    if not homework_id:
        return jsonify({'success': False, 'message': '作业ID不能为空'}), 400
    
    # 检查学生是否存在
    student = Student.query.get(student_id)
    if not student:
        return jsonify({'success': False, 'message': '学生不存在'}), 404
    
    # 检查作业是否存在
    homework = Homework.query.get(homework_id)
    if not homework:
        return jsonify({'success': False, 'message': '作业不存在'}), 404
    
    # 检查是否已提交
    existing_submission = HomeworkSubmission.query.filter_by(
        student_id=student_id,
        homework_id=homework_id
    ).first()
    if existing_submission:
        return jsonify({
            'success': True,
            'message': '提交记录已存在',
            'submission_id': existing_submission.id,
            'subject': homework.subject,
            'max_images': homework.max_images or 5
        }), 200
    
    try:
        # 创建提交记录
        submission = HomeworkSubmission(student_id=student_id, homework_id=homework_id)
        db.session.add(submission)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': '请开始拍照上传作业',
            'submission_id': submission.id,
            'subject': homework.subject,
            'max_images': homework.max_images or 5
        }), 200
    except Exception as e:
        db.session.rollback()
        print(f"创建提交记录失败: {str(e)}")
        return jsonify({'success': False, 'message': '创建失败,请重试'}), 500

def get_ai_session_cookie():
    """获取或刷新AI API的session cookie"""
    global ai_session_cookie
    
    with ai_cookie_lock:
        # 如果已有cookie，先尝试使用
        if ai_session_cookie:
            return ai_session_cookie
        
        # 登录获取新cookie
        try:
            response = requests.post(
                AI_LOGIN_URL,
                json={
                    'username': AI_USERNAME,
                    'password': AI_PASSWORD
                },
                timeout=10
            )
            
            if response.status_code == 200:
                # 从响应头中提取session cookie
                set_cookie_header = response.headers.get('Set-Cookie', '')
                if 'session=' in set_cookie_header:
                    # 提取session值
                    session_match = re.search(r'session=([^;]+)', set_cookie_header)
                    if session_match:
                        ai_session_cookie = session_match.group(1)
                        print(f"[AI] 成功获取session cookie")
                        return ai_session_cookie
            
            print(f"[AI] 登录失败: {response.status_code}")
            return None
            
        except Exception as e:
            print(f"[AI] 登录异常: {str(e)}")
            return None

def call_ai_review(submission_id):
    """调用AI进行作业审核（异步执行）"""
    global ai_session_cookie
    
    with app.app_context():
        try:
            submission = HomeworkSubmission.query.get(submission_id)
            if not submission:
                return
            
            # 设置为"判定中"状态
            submission.ai_review_status = 'reviewing'
            submission.ai_review_result = 'AI正在判定中...'
            db.session.commit()
            
            # 获取该提交的所有图片
            images = HomeworkImage.query.filter_by(submission_id=submission_id).all()
            if not images:
                submission.ai_review_status = 'approved'
                submission.ai_review_result = '无图片，自动通过'
                submission.ai_reviewed_at = get_china_time()
                db.session.commit()
                return
            
            # 获取作业信息，使用自定义prompt
            homework = submission.homework
            custom_prompt = homework.ai_prompt
            
            # 构建图片URL列表
            image_urls = []
            for img in images:
                image_url = f"{HOMEWORK_BASE_URL}/uploads/{img.filename}"
                print(image_url)
                image_urls.append({
                    "type": "image_url",
                    "image_url": {"url": image_url}
                })
            
            # 构建消息内容 - 使用自定义prompt或默认prompt
            if custom_prompt:
                prompt_text = f"{custom_prompt}\n\n请严格按照以下JSON格式输出，不要添加任何其他文字或解释：\n{{\"ok\": true}}  或  {{\"ok\": false}}\n\n其中ok为true表示这些图片符合要求，ok为false表示不符合要求。"
            else:
                prompt_text = "请仔细查看这些图片，判断它们是否看起来像是学生提交的作业（例如：作业本、试卷、练习题、手写内容等）。\n\n请严格按照以下JSON格式输出，不要添加任何其他文字或解释：\n{\"ok\": true}  或  {\"ok\": false}\n\n其中ok为true表示这些图片看起来像作业，ok为false表示不像作业。"
            
            content = [{
                "type": "text",
                "text": prompt_text
            }]
            content.extend(image_urls)
            
            # 准备API请求
            payload = {
                "model": AI_MODEL,
                "group": "default",
                "messages": [
                    {
                        "role": "system",
                        "content": "你是一个作业审核助手。你的任务是判断图片是否为学生作业。请只输出JSON格式的结果，不要添加任何其他内容。"
                    },
                    {
                        "role": "user",
                        "content": content
                    }
                ],
                "stream": True,
                "temperature": 0.3,
                "top_p": 1,
                "frequency_penalty": 0,
                "presence_penalty": 0
            }
            
            # 尝试多次调用AI API
            for attempt in range(AI_REVIEW_MAX_RETRIES):
                try:
                    print(f"[AI] 开始第 {attempt + 1}/{AI_REVIEW_MAX_RETRIES} 次审核尝试 - Submission ID: {submission_id}")
                    
                    # 获取session cookie
                    session_cookie = get_ai_session_cookie()
                    if not session_cookie:
                        print(f"[AI] 无法获取session cookie，跳过审核")
                        continue
                    
                    # 准备请求头
                    headers = {
                        'Content-Type': 'application/json',
                        'Cookie': f'session={session_cookie}',
                        'New-Api-User': '2'
                    }
                    
                    print(f"[AI] 发送API请求到: {AI_API_URL}")
                    print(f"[AI] 使用模型: {AI_MODEL}")
                    print(f"[AI] 图片数量: {len(images)}")
                    
                    response = requests.post(
                        AI_API_URL,
                        json=payload,
                        headers=headers,
                        stream=True,
                        timeout=60
                    )
                    
                    print(f"[AI] API响应状态码: {response.status_code}")
                    
                    print(f"[AI] API响应状态码: {response.status_code}")
                    try:
                    # 只在非 200 时打印完整响应体，避免日志太大
                        if response.status_code != 200:
                            print(f"[AI] API响应体（非200）: {response.text}")  # 截断避免过长
                    except Exception as log_e:
                        print(f"[AI] 打印响应体时出错: {log_e}")
                    
                    # 解析流式响应
                    full_content = ""
                    for line in response.iter_lines():
                        if line:
                            line_text = line.decode('utf-8')
                            if line_text.startswith('data: '):
                                line_text = line_text[6:]
                            
                            if line_text.strip() == '[DONE]':
                                break
                            
                            try:
                                chunk = json.loads(line_text)
                                if 'choices' in chunk and len(chunk['choices']) > 0:
                                    delta = chunk['choices'][0].get('delta', {})
                                    if 'content' in delta:
                                        full_content += delta['content']
                            except json.JSONDecodeError:
                                continue
                    
                    # 处理返回的内容，提取JSON
                    # 移除可能的代码块标记
                    content_clean = full_content.strip()
                    content_clean = re.sub(r'^```json\s*', '', content_clean)
                    content_clean = re.sub(r'^```\s*', '', content_clean)
                    content_clean = re.sub(r'\s*```$', '', content_clean)
                    content_clean = content_clean.strip()
                    
                    # 尝试解析JSON
                    try:
                        result = json.loads(content_clean)
                        print(f"[AI] 成功解析AI响应: {result}")
                        
                        if 'ok' in result and isinstance(result['ok'], bool):
                            # 成功解析，更新数据库
                            if result['ok']:
                                print(f"[AI] ✓ AI判定为正常作业 - Submission ID: {submission_id}")
                                submission.ai_review_status = 'approved'
                                submission.ai_review_result = '通过AI审核'
                            else:
                                print(f"[AI] ✗ AI判定为异常作业 - Submission ID: {submission_id}")
                                # AI判定不像作业，根据配置处理
                                if AI_REVIEW_ACTION == 'reject':
                                    print(f"[AI] 执行操作: 打回作业并删除记录")
                                    # 打回作业 - 删除提交记录和图片
                                    submission.ai_review_status = 'rejected'
                                    submission.ai_review_result = 'AI判定不像作业，已自动打回'
                                    submission.ai_reviewed_at = get_china_time()
                                    db.session.commit()
                                    
                                    # 删除图片和提交记录
                                    for img in images:
                                        filepath = os.path.join(app.config['UPLOAD_FOLDER'], img.filename)
                                        if os.path.exists(filepath):
                                            os.remove(filepath)
                                        db.session.delete(img)
                                    db.session.delete(submission)
                                    db.session.commit()
                                    return
                                    
                                elif AI_REVIEW_ACTION == 'mark_abnormal':
                                    print(f"[AI] 执行操作: 标记为异常，保留记录")
                                    # 标记为异常，保留提交记录
                                    submission.ai_review_status = 'rejected'
                                    submission.ai_review_result = 'AI判定不像作业，已标记为异常'
                                    
                                else:  # ignore
                                    print(f"[AI] 执行操作: 忽略AI判断")
                                    # 忽略AI判断，标记但不影响提交
                                    submission.ai_review_status = 'rejected'
                                    submission.ai_review_result = 'AI判定不像作业（已忽略）'

                            submission.ai_reviewed_at = get_china_time()
                            db.session.commit()
                            print(f"[AI] 审核完成 - Submission ID: {submission_id}, 状态: {submission.ai_review_status}")
                            return
                    except json.JSONDecodeError as je:
                        print(f"[AI] JSON解析失败: {str(je)}")
                        print(f"[AI] 原始内容: {content_clean[:200]}...")
                        # JSON解析失败，继续重试
                        continue
                        
                except Exception as e:
                    print(f"[AI] ✗ 审核尝试 {attempt + 1} 失败: {str(e)}")
                    import traceback
                    traceback.print_exc()
                    continue
            
            # 所有重试都失败
            print(f"[AI] ✗ 所有重试均失败 - Submission ID: {submission_id}")
            submission.ai_review_status = 'error'
            submission.ai_review_result = 'AI审核失败，已达最大重试次数'
            submission.ai_reviewed_at = get_china_time()
            db.session.commit()
            
        except Exception as e:
            print(f"[AI] ✗ 审核异常 - Submission ID: {submission_id}")
            print(f"[AI] 异常详情: {str(e)}")
            import traceback
            traceback.print_exc()
            try:
                submission = HomeworkSubmission.query.get(submission_id)
                if submission:
                    submission.ai_review_status = 'error'
                    submission.ai_review_result = f'审核异常: {str(e)}'
                    submission.ai_reviewed_at = get_china_time()
                    db.session.commit()
            except:
                pass

@app.route('/api/confirm-submission/<int:submission_id>', methods=['POST'])
def confirm_submission(submission_id):
    """确认提交作业（拍照后或直接提交）"""
    submission = HomeworkSubmission.query.get(submission_id)
    if not submission:
        return jsonify({'success': False, 'message': '提交记录不存在'}), 404

    # 只有在启用了图片上传功能时，才检查是否至少上传了一张图片
    if ENABLE_IMAGE_UPLOAD:
        image_count = HomeworkImage.query.filter_by(submission_id=submission_id).count()
        if image_count == 0:
            return jsonify({'success': False, 'message': '请至少上传一张作业图片'}), 400

    homework = submission.homework
    teacher = homework.teacher  # 获取作业对应的教师

    # 检查全局AI审核开关和教师个人AI审核开关
    ai_review_enabled = ENABLE_AI_REVIEW and teacher.enable_ai_review and ENABLE_IMAGE_UPLOAD

    # 如果启用了AI审核且有图片，启动异步审核
    if ai_review_enabled:
        submission.ai_review_status = 'reviewing'
        submission.ai_review_result = 'AI正在判定中...'
        db.session.commit()

        # 在后台线程中执行AI审核
        thread = threading.Thread(target=call_ai_review, args=(submission_id,))
        thread.daemon = True
        thread.start()

    return jsonify({
        'success': True,
        'message': f'{homework.subject}作业提交成功！' + ('正在进行AI审核...' if ai_review_enabled else ''),
        'submitted_at': submission.submitted_at.strftime('%Y-%m-%d %H:%M:%S'),
        'ai_review_enabled': ai_review_enabled
    })

@app.route('/api/upload-image', methods=['POST'])
def upload_image():
    """上传作业图片（Base64格式）"""
    if not ENABLE_IMAGE_UPLOAD:
        return jsonify({'success': False, 'message': '图片上传功能未启用'}), 403
    
    data = request.get_json()
    submission_id = data.get('submission_id')
    image_data = data.get('image_data')  # Base64编码的图片数据
    
    if not submission_id:
        return jsonify({'success': False, 'message': '提交记录ID不能为空'}), 400
    
    if not image_data:
        return jsonify({'success': False, 'message': '图片数据不能为空'}), 400
    
    # 检查提交记录是否存在
    submission = HomeworkSubmission.query.get(submission_id)
    if not submission:
        return jsonify({'success': False, 'message': '提交记录不存在'}), 404
    
    # 检查图片数量限制 - 使用作业的max_images设置
    homework = submission.homework
    max_images = homework.max_images or MAX_IMAGES_PER_HOMEWORK
    current_image_count = HomeworkImage.query.filter_by(submission_id=submission_id).count()
    if current_image_count >= max_images:
        return jsonify({'success': False, 'message': f'最多只能上传{max_images}张图片'}), 400
    
    try:
        import base64
        import io
        from PIL import Image
        
        # 解析Base64数据
        if ',' in image_data:
            image_data = image_data.split(',')[1]
        
        # 解码Base64
        image_bytes = base64.b64decode(image_data)
        
        # 打开图片并转换为JPEG
        image = Image.open(io.BytesIO(image_bytes))
        
        # 生成唯一文件名
        filename = f"{uuid.uuid4().hex}.jpg"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        
        # 保存为JPEG格式
        if image.mode in ('RGBA', 'LA', 'P'):
            image = image.convert('RGB')
        image.save(filepath, 'JPEG', quality=85)
        
        # 保存到数据库
        db_image = HomeworkImage(
            submission_id=submission_id,
            filename=filename,
            original_filename=f"camera_{datetime.now().strftime('%Y%m%d_%H%M%S')}.jpg"
        )
        db.session.add(db_image)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': '图片上传成功',
            'image': {
                'id': db_image.id,
                'filename': filename,
                'uploaded_at': db_image.uploaded_at.strftime('%Y-%m-%d %H:%M:%S')
            }
        }), 200
    except Exception as e:
        db.session.rollback()
        print(f"上传图片失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': '上传失败,请重试'}), 500

@app.route('/api/submission-images/<int:submission_id>')
def get_submission_images(submission_id):
    """获取提交记录的所有图片"""
    submission = HomeworkSubmission.query.get(submission_id)
    if not submission:
        return jsonify({'success': False, 'message': '提交记录不存在'}), 404
    
    images = HomeworkImage.query.filter_by(submission_id=submission_id).order_by(HomeworkImage.uploaded_at).all()
    
    image_list = []
    for img in images:
        image_list.append({
            'id': img.id,
            'filename': img.filename,
            'original_filename': img.original_filename,
            'uploaded_at': img.uploaded_at.strftime('%Y-%m-%d %H:%M:%S'),
            'url': f'/uploads/{img.filename}'
        })
    
    return jsonify(image_list)

@app.route('/api/delete-image/<int:image_id>', methods=['DELETE'])
def delete_image(image_id):
    """删除图片"""
    if not ENABLE_IMAGE_UPLOAD:
        return jsonify({'success': False, 'message': '图片上传功能未启用'}), 403
    
    image = HomeworkImage.query.get(image_id)
    if not image:
        return jsonify({'success': False, 'message': '图片不存在'}), 404
    
    try:
        # 删除文件
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], image.filename)
        if os.path.exists(filepath):
            os.remove(filepath)
        
        # 删除数据库记录
        db.session.delete(image)
        db.session.commit()
        
        return jsonify({'success': True, 'message': '图片删除成功'}), 200
    except Exception as e:
        db.session.rollback()
        print(f"删除图片失败: {str(e)}")
        return jsonify({'success': False, 'message': '删除失败,请重试'}), 500

@app.route('/uploads/<filename>')
def uploaded_file(filename):
    """访问上传的图片"""
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)

@app.route('/api/delete-submission/<int:submission_id>', methods=['DELETE'])
def delete_submission(submission_id):
    """删除作业提交记录（用于重新提交）"""
    submission = HomeworkSubmission.query.get(submission_id)
    if not submission:
        return jsonify({'success': False, 'message': '提交记录不存在'}), 404
    
    try:
        # 删除相关图片
        images = HomeworkImage.query.filter_by(submission_id=submission_id).all()
        for img in images:
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], img.filename)
            if os.path.exists(filepath):
                os.remove(filepath)
            db.session.delete(img)
        
        # 删除提交记录
        db.session.delete(submission)
        db.session.commit()
        
        return jsonify({'success': True, 'message': '提交记录已删除'}), 200
    except Exception as e:
        db.session.rollback()
        print(f"删除提交记录失败: {str(e)}")
        return jsonify({'success': False, 'message': '删除失败,请重试'}), 500

# ==================== 教师端路由 ====================
@app.route('/teacher')
def teacher_index():
    """教师端首页"""
    if 'teacher_id' not in session:
        return redirect(url_for('teacher_login'))
    return render_template('teacher.html', teacher_subject=session.get('teacher_subject'))

@app.route('/teacher/login')
def teacher_login():
    """教师登录页面"""
    return render_template('teacher_login.html')

@app.route('/api/teacher/login', methods=['POST'])
def api_teacher_login():
    """教师登录接口"""
    data = request.get_json()
    username = data.get('username')
    password = data.get('password')
    
    if not username or not password:
        return jsonify({'success': False, 'message': '用户名和密码不能为空'}), 400
    
    teacher = Teacher.query.filter_by(username=username).first()
    if not teacher or not check_password_hash(teacher.password, password):
        return jsonify({'success': False, 'message': '用户名或密码错误'}), 401
    
    session['teacher_id'] = teacher.id
    session['teacher_username'] = teacher.username
    session['teacher_subject'] = teacher.subject
    
    return jsonify({'success': True, 'message': '登录成功'}), 200

@app.route('/api/teacher/register', methods=['POST'])
def api_teacher_register():
    """教师注册接口 - 已禁用，只能通过管理员添加"""
    return jsonify({'success': False, 'message': '教师注册功能已关闭，请联系管理员添加账户'}), 403

@app.route('/api/teacher/logout', methods=['POST'])
def api_teacher_logout():
    """教师登出接口"""
    session.pop('teacher_id', None)
    session.pop('teacher_username', None)
    session.pop('teacher_subject', None)
    return jsonify({'success': True, 'message': '登出成功'}), 200

@app.route('/api/teacher/current-info')
def get_teacher_info():
    """获取当前教师信息"""
    if 'teacher_id' not in session:
        return jsonify({'success': False, 'message': '未登录'}), 401

    teacher = Teacher.query.get(session.get('teacher_id'))
    if not teacher:
        return jsonify({'success': False, 'message': '教师不存在'}), 404

    return jsonify({
        'success': True,
        'username': session.get('teacher_username'),
        'subject': session.get('teacher_subject'),
        'enable_ai_review': teacher.enable_ai_review
    })

@app.route('/api/teacher/add-student', methods=['POST'])
def api_add_student():
    """添加学生 - 已禁用，只能通过管理员添加"""
    return jsonify({'success': False, 'message': '此功能已禁用，请使用管理端添加学生'}), 403

@app.route('/api/teacher/unsubmitted-students')
def get_unsubmitted_students():
    """获取未提交作业的学生名单"""
    if 'teacher_id' not in session:
        return jsonify({'success': False, 'message': '未登录'}), 401
    
    teacher_id = session.get('teacher_id')
    
    # 获取该教师布置的所有作业
    homeworks = Homework.query.filter_by(teacher_id=teacher_id).all()
    if not homeworks:
        return jsonify([])
    
    # 获取所有学生
    all_students = Student.query.all()
    unsubmitted_students = []
    
    for student in all_students:
        has_unsubmitted = False
        for hw in homeworks:
            submission = HomeworkSubmission.query.filter_by(
                student_id=student.id,
                homework_id=hw.id
            ).first()
            if not submission:
                has_unsubmitted = True
                break
        
        if has_unsubmitted:
            unsubmitted_students.append({
                'id': student.id,
                'name': student.name,
                'student_id': student.student_id
            })
    
    return jsonify(unsubmitted_students)

@app.route('/api/teacher/all-students-status')
def get_all_students_status():
    """获取所有学生及作业提交状态"""
    if 'teacher_id' not in session:
        return jsonify({'success': False, 'message': '未登录'}), 401
    
    teacher_id = session.get('teacher_id')
    students = Student.query.all()
    # 获取该教师布置的所有作业
    homeworks = Homework.query.filter_by(teacher_id=teacher_id).all()
    
    student_list = []
    for student in students:
        homework_details = []
        submitted_count = 0
        
        for hw in homeworks:
            submission = HomeworkSubmission.query.filter_by(
                student_id=student.id,
                homework_id=hw.id
            ).first()
            
            if submission:
                submitted_count += 1
            
            # 获取图片数量和AI审核状态
            image_count = 0
            ai_review_status = None
            ai_review_result = None
            if submission:
                image_count = HomeworkImage.query.filter_by(submission_id=submission.id).count()
                ai_review_status = submission.ai_review_status
                ai_review_result = submission.ai_review_result
            
            homework_details.append({
                'homework_id': hw.id,
                'title': hw.title,
                'subject': hw.subject,
                'submitted': submission is not None,
                'submitted_at': submission.submitted_at.strftime('%Y-%m-%d %H:%M:%S') if submission else None,
                'submission_id': submission.id if submission else None,
                'image_count': image_count,
                'ai_review_status': ai_review_status,
                'ai_review_result': ai_review_result
            })
        
        student_list.append({
            'id': student.id,
            'name': student.name,
            'student_id': student.student_id,
            'homework_details': homework_details,
            'submitted_count': submitted_count,
            'total_homework': len(homeworks)
        })
    
    return jsonify(student_list)

@app.route('/api/teacher/reset-submissions', methods=['POST'])
def reset_submissions():
    """还原作业提交情况(清空该教师布置作业的提交记录)"""
    if 'teacher_id' not in session:
        return jsonify({'success': False, 'message': '未登录'}), 401
    
    teacher_id = session.get('teacher_id')
    
    try:
        # 获取该教师布置的所有作业
        homeworks = Homework.query.filter_by(teacher_id=teacher_id).all()
        homework_ids = [hw.id for hw in homeworks]
        
        # 删除这些作业的所有提交记录
        HomeworkSubmission.query.filter(HomeworkSubmission.homework_id.in_(homework_ids)).delete(synchronize_session=False)
        db.session.commit()
        
        return jsonify({'success': True, 'message': '作业提交记录已还原'}), 200
    except Exception as e:
        db.session.rollback()
        print(f"还原失败: {str(e)}")
        return jsonify({'success': False, 'message': '操作失败,请重试'}), 500

@app.route('/api/teacher/delete-student/<int:student_id>', methods=['DELETE'])
def delete_student(student_id):
    """删除学生 - 已禁用，只能通过管理员删除"""
    return jsonify({'success': False, 'message': '此功能已禁用，请使用管理端删除学生'}), 403

@app.route('/api/teacher/publish-homework', methods=['POST'])
def publish_homework():
    """教师布置作业"""
    if 'teacher_id' not in session:
        return jsonify({'success': False, 'message': '未登录'}), 401
    
    data = request.get_json()
    title = data.get('title')
    ai_prompt = data.get('ai_prompt', '')  # 自定义AI prompt
    max_images = data.get('max_images', 5)  # 最大图片数量，默认5
    teacher_id = session.get('teacher_id')
    teacher_subject = session.get('teacher_subject')
    
    if not title:
        return jsonify({'success': False, 'message': '作业标题不能为空'}), 400
    
    # 验证max_images
    try:
        max_images = int(max_images)
        if max_images < 1 or max_images > 20:
            return jsonify({'success': False, 'message': '图片数量必须在1-20之间'}), 400
    except (ValueError, TypeError):
        max_images = 5
    
    try:
        homework = Homework(
            title=title,
            subject=teacher_subject,
            teacher_id=teacher_id,
            ai_prompt=ai_prompt if ai_prompt else None,
            max_images=max_images
        )
        db.session.add(homework)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': '作业布置成功',
            'homework': {
                'id': homework.id,
                'title': homework.title,
                'subject': homework.subject,
                'ai_prompt': homework.ai_prompt,
                'max_images': homework.max_images,
                'created_at': homework.created_at.strftime('%Y-%m-%d %H:%M:%S')
            }
        }), 200
    except Exception as e:
        db.session.rollback()
        print(f"布置作业失败: {str(e)}")
        return jsonify({'success': False, 'message': '布置失败,请重试'}), 500

@app.route('/api/teacher/homeworks')
def get_teacher_homeworks():
    """获取教师布置的所有作业"""
    if 'teacher_id' not in session:
        return jsonify({'success': False, 'message': '未登录'}), 401
    
    teacher_id = session.get('teacher_id')
    homeworks = Homework.query.filter_by(teacher_id=teacher_id).order_by(Homework.created_at.desc()).all()
    
    homework_list = []
    for hw in homeworks:
        # 统计提交情况
        total_students = Student.query.count()
        submitted_count = HomeworkSubmission.query.filter_by(homework_id=hw.id).count()
        
        homework_list.append({
            'id': hw.id,
            'title': hw.title,
            'subject': hw.subject,
            'created_at': hw.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            'total_students': total_students,
            'submitted_count': submitted_count
        })
    
    return jsonify(homework_list)

@app.route('/api/teacher/delete-homework/<int:homework_id>', methods=['DELETE'])
def delete_homework(homework_id):
    """删除作业"""
    if 'teacher_id' not in session:
        return jsonify({'success': False, 'message': '未登录'}), 401
    
    teacher_id = session.get('teacher_id')
    homework = Homework.query.get(homework_id)
    
    if not homework:
        return jsonify({'success': False, 'message': '作业不存在'}), 404
    
    if homework.teacher_id != teacher_id:
        return jsonify({'success': False, 'message': '无权限删除此作业'}), 403
    
    try:
        # 删除相关图片
        submissions = HomeworkSubmission.query.filter_by(homework_id=homework_id).all()
        for sub in submissions:
            images = HomeworkImage.query.filter_by(submission_id=sub.id).all()
            for img in images:
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], img.filename)
                if os.path.exists(filepath):
                    os.remove(filepath)
                db.session.delete(img)
        
        # 删除相关提交记录
        HomeworkSubmission.query.filter_by(homework_id=homework_id).delete()
        # 删除作业
        db.session.delete(homework)
        db.session.commit()
        
        return jsonify({'success': True, 'message': '作业删除成功'}), 200
    except Exception as e:
        db.session.rollback()
        print(f"删除作业失败: {str(e)}")
        return jsonify({'success': False, 'message': '删除失败,请重试'}), 500

@app.route('/api/teacher/override-ai-review/<int:submission_id>', methods=['POST'])
def override_ai_review(submission_id):
    """教师手动覆盖AI审核结果"""
    if 'teacher_id' not in session:
        return jsonify({'success': False, 'message': '未登录'}), 401

    data = request.get_json()
    action = data.get('action')  # approve, reject_and_delete

    if action not in ['approve', 'reject_and_delete']:
        return jsonify({'success': False, 'message': '无效的操作'}), 400

    submission = HomeworkSubmission.query.get(submission_id)
    if not submission:
        return jsonify({'success': False, 'message': '提交记录不存在'}), 404

    try:
        if action == 'approve':
            # 教师批准，覆盖AI判断
            submission.ai_review_status = 'approved'
            submission.ai_review_result = '教师手动批准'
            submission.ai_reviewed_at = get_china_time()
            db.session.commit()
            return jsonify({'success': True, 'message': '已批准该作业'}), 200

        elif action == 'reject_and_delete':
            # 教师确认打回，删除提交记录
            images = HomeworkImage.query.filter_by(submission_id=submission_id).all()
            for img in images:
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], img.filename)
                if os.path.exists(filepath):
                    os.remove(filepath)
                db.session.delete(img)

            db.session.delete(submission)
            db.session.commit()
            return jsonify({'success': True, 'message': '已打回该作业'}), 200

    except Exception as e:
        db.session.rollback()
        print(f"覆盖AI审核失败: {str(e)}")
        return jsonify({'success': False, 'message': '操作失败，请重试'}), 500

@app.route('/api/teacher/retry-ai-review/<int:submission_id>', methods=['POST'])
def retry_ai_review(submission_id):
    """手动重试AI审核"""
    if 'teacher_id' not in session:
        return jsonify({'success': False, 'message': '未登录'}), 401

    submission = HomeworkSubmission.query.get(submission_id)
    if not submission:
        return jsonify({'success': False, 'message': '提交记录不存在'}), 404

    # 只有在状态为 error 或 rejected 时才允许重试
    if submission.ai_review_status not in ['error', 'rejected']:
        return jsonify({'success': False, 'message': '只有AI判定异常或审核失败的作业才能重试'}), 400

    try:
        # 更新状态为 reviewing（判定中）
        submission.ai_review_status = 'reviewing'
        submission.ai_review_result = 'AI正在重新判定中...'
        db.session.commit()

        # 启动后台线程进行审核
        thread = threading.Thread(target=call_ai_review, args=(submission_id,))
        thread.daemon = True
        thread.start()

        return jsonify({'success': True, 'message': '已启动AI重审，请稍候...'}), 200

    except Exception as e:
        db.session.rollback()
        print(f"重试AI审核失败: {str(e)}")
        return jsonify({'success': False, 'message': '操作失败，请重试'}), 500

@app.route('/api/teacher/daily-homework-stats')
def get_daily_homework_stats():
    """获取指定日期布置的作业及提交统计"""
    if 'teacher_id' not in session:
        return jsonify({'success': False, 'message': '未登录'}), 401
    
    teacher_id = session.get('teacher_id')
    date_str = request.args.get('date')  # 格式: YYYY-MM-DD
    
    if not date_str:
        return jsonify({'success': False, 'message': '请提供日期参数'}), 400
    
    try:
        # 解析日期
        date_obj = datetime.strptime(date_str, '%Y-%m-%d').replace(tzinfo=CHINA_TZ)
        date_start = date_obj.replace(hour=0, minute=0, second=0, microsecond=0)
        date_end = date_obj.replace(hour=23, minute=59, second=59, microsecond=999999)
        
        # 获取该日期布置的所有作业
        daily_homeworks = Homework.query.filter(
            Homework.teacher_id == teacher_id,
            Homework.created_at >= date_start,
            Homework.created_at <= date_end
        ).all()
        
        if not daily_homeworks:
            return jsonify({
                'date': date_str,
                'homeworks': [],
                'summary': {
                    'total': 0,
                    'submitted': 0,
                    'ai_rejected': 0,
                    'ai_error': 0,
                    'not_submitted': 0,
                    'submitted_percent': 0,
                    'ai_rejected_percent': 0,
                    'ai_error_percent': 0,
                    'not_submitted_percent': 0
                }
            })
        
        # 获取所有学生
        total_students = Student.query.count()
        
        # 统计每个作业的提交情况
        homework_stats = []
        total_submitted = 0
        total_ai_rejected = 0
        total_ai_error = 0
        
        for hw in daily_homeworks:
            # 获取该作业的所有提交记录
            submissions = HomeworkSubmission.query.filter_by(homework_id=hw.id).all()
            submitted_count = len(submissions)
            ai_rejected_count = sum(1 for s in submissions if s.ai_review_status == 'rejected')
            ai_error_count = sum(1 for s in submissions if s.ai_review_status == 'error')
            
            total_submitted += submitted_count
            total_ai_rejected += ai_rejected_count
            total_ai_error += ai_error_count
            
            homework_stats.append({
                'homework_id': hw.id,
                'title': hw.title,
                'subject': hw.subject,
                'created_at': hw.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                'total_students': total_students,
                'submitted': submitted_count,
                'ai_rejected': ai_rejected_count,
                'ai_error': ai_error_count,
                'not_submitted': total_students - submitted_count
            })
        
        # 计算总体统计
        total_assignments = len(daily_homeworks) * total_students
        total_not_submitted = total_assignments - total_submitted
        total_abnormal = total_ai_rejected + total_ai_error  # 异常总数
        
        summary = {
            'total': total_assignments,
            'submitted': total_submitted - total_abnormal,  # 正常提交（已提交 - 异常）
            'ai_rejected': total_ai_rejected,  # AI判定不像作业
            'ai_error': total_ai_error,  # AI审核异常
            'not_submitted': total_not_submitted,
            'submitted_percent': round((total_submitted - total_abnormal) / total_assignments * 100, 1) if total_assignments > 0 else 0,
            'ai_rejected_percent': round(total_ai_rejected / total_assignments * 100, 1) if total_assignments > 0 else 0,
            'ai_error_percent': round(total_ai_error / total_assignments * 100, 1) if total_assignments > 0 else 0,
            'not_submitted_percent': round(total_not_submitted / total_assignments * 100, 1) if total_assignments > 0 else 0
        }
        
        return jsonify({
            'date': date_str,
            'homeworks': homework_stats,
            'summary': summary
        })
        
    except ValueError:
        return jsonify({'success': False, 'message': '日期格式错误，请使用YYYY-MM-DD格式'}), 400
    except Exception as e:
        print(f"获取每日作业统计失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': '获取统计数据失败'}), 500

@app.route('/api/teacher/homework-dates')
def get_homework_dates():
    """获取有作业布置的所有日期列表"""
    if 'teacher_id' not in session:
        return jsonify({'success': False, 'message': '未登录'}), 401
    
    teacher_id = session.get('teacher_id')
    
    try:
        # 获取该教师布置的所有作业
        homeworks = Homework.query.filter_by(teacher_id=teacher_id).order_by(Homework.created_at.desc()).all()
        
        # 提取唯一的日期
        dates = list(set([hw.created_at.strftime('%Y-%m-%d') for hw in homeworks]))
        dates.sort(reverse=True)  # 降序排列
        
        return jsonify(dates)
        
    except Exception as e:
        print(f"获取作业日期列表失败: {str(e)}")
        return jsonify({'success': False, 'message': '获取日期列表失败'}), 500

@app.route('/api/teacher/abnormal-submissions')
def get_abnormal_submissions():
    """获取异常作业列表（AI审核未通过的作业，包括判定中）"""
    if 'teacher_id' not in session:
        return jsonify({'success': False, 'message': '未登录'}), 401
    
    teacher_id = session.get('teacher_id')
    
    try:
        # 获取该教师布置的所有作业
        homeworks = Homework.query.filter_by(teacher_id=teacher_id).all()
        homework_ids = [hw.id for hw in homeworks]
        
        if not homework_ids:
            return jsonify([])
        
        # 获取AI审核未通过的提交记录（包括判定中、已拒绝、审核失败）
        abnormal_submissions = HomeworkSubmission.query.filter(
            HomeworkSubmission.homework_id.in_(homework_ids),
            HomeworkSubmission.ai_review_status.in_(['reviewing', 'rejected', 'error'])
        ).order_by(HomeworkSubmission.submitted_at.desc()).all()
        
        result = []
        for submission in abnormal_submissions:
            student = submission.student
            homework = submission.homework
            image_count = HomeworkImage.query.filter_by(submission_id=submission.id).count()
            
            result.append({
                'submission_id': submission.id,
                'student_name': student.name,
                'student_id': student.student_id,
                'homework_title': homework.title,
                'homework_subject': homework.subject,
                'submitted_at': submission.submitted_at.strftime('%Y-%m-%d %H:%M:%S'),
                'ai_review_status': submission.ai_review_status,
                'ai_review_result': submission.ai_review_result,
                'image_count': image_count
            })
        
        return jsonify(result)
        
    except Exception as e:
        print(f"获取异常作业列表失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': '获取异常作业列表失败'}), 500

@app.route('/api/teacher/toggle-ai-review', methods=['POST'])
def toggle_ai_review():
    """切换教师的AI复审设置"""
    if 'teacher_id' not in session:
        return jsonify({'success': False, 'message': '未登录'}), 401

    teacher_id = session.get('teacher_id')
    teacher = Teacher.query.get(teacher_id)

    if not teacher:
        return jsonify({'success': False, 'message': '教师不存在'}), 404

    try:
        # 切换AI复审设置
        teacher.enable_ai_review = not teacher.enable_ai_review
        db.session.commit()

        status = "启用" if teacher.enable_ai_review else "禁用"
        return jsonify({
            'success': True,
            'message': f'AI复审已{status}',
            'enable_ai_review': teacher.enable_ai_review
        }), 200
    except Exception as e:
        db.session.rollback()
        print(f"切换AI复审设置失败: {str(e)}")
        return jsonify({'success': False, 'message': '操作失败，请重试'}), 500

@app.route('/api/teacher/export-homework/<int:homework_id>')
def export_homework_submissions(homework_id):
    """导出某项作业的提交情况"""
    if 'teacher_id' not in session:
        return jsonify({'success': False, 'message': '未登录'}), 401
    
    format_type = request.args.get('format', 'excel')  # excel, csv, json, txt
    
    homework = Homework.query.get(homework_id)
    if not homework:
        return jsonify({'success': False, 'message': '作业不存在'}), 404
    
    if homework.teacher_id != session.get('teacher_id'):
        return jsonify({'success': False, 'message': '无权限导出此作业'}), 403
    
    # 获取所有学生及其提交状态
    students = Student.query.all()
    data = []
    
    for student in students:
        submission = HomeworkSubmission.query.filter_by(
            student_id=student.id,
            homework_id=homework_id
        ).first()
        
        if submission:
            status = '已提交'
            if submission.ai_review_status == 'approved':
                status = '已提交-AI审核通过'
            elif submission.ai_review_status == 'rejected':
                status = '已提交-AI判定异常'
            elif submission.ai_review_status == 'reviewing':
                status = '已提交-AI判定中'
            elif submission.ai_review_status == 'error':
                status = '已提交-AI审核失败'
        else:
            status = '未提交'
        
        data.append({
            '学号': student.student_id,
            '姓名': student.name,
            '状态': status
        })
    
    return generate_export_file(data, f"{homework.subject}_{homework.title}_提交情况", format_type)

@app.route('/api/teacher/export-student/<int:student_id>')
def export_student_submissions(student_id):
    """导出某个学生的历史作业提交记录"""
    if 'teacher_id' not in session:
        return jsonify({'success': False, 'message': '未登录'}), 401
    
    format_type = request.args.get('format', 'excel')
    teacher_id = session.get('teacher_id')
    
    student = db.session.get(Student, student_id)
    if not student:
        return jsonify({'success': False, 'message': '学生不存在'}), 404
    
    # 获取该教师布置的所有作业
    homeworks = Homework.query.filter_by(teacher_id=teacher_id).all()
    data = []
    
    for hw in homeworks:
        submission = HomeworkSubmission.query.filter_by(
            student_id=student_id,
            homework_id=hw.id
        ).first()
        
        if submission:
            status = '已提交'
            if submission.ai_review_status == 'approved':
                status = '已提交-AI审核通过'
            elif submission.ai_review_status == 'rejected':
                status = '已提交-AI判定异常'
            elif submission.ai_review_status == 'reviewing':
                status = '已提交-AI判定中'
            elif submission.ai_review_status == 'error':
                status = '已提交-AI审核失败'
        else:
            status = '未提交'
        
        data.append({
            '作业名': hw.title,
            '作业学科': hw.subject,
            '布置时间': hw.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            '状态': status
        })
    
    return generate_export_file(data, f"{student.name}_{student.student_id}_作业记录", format_type)

def generate_export_file(data, filename, format_type):
    """生成导出文件"""
    import io
    import csv
    import json
    from datetime import datetime
    
    # 使用时间戳作为文件名，避免中文编码问题
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    safe_filename = f"export_{timestamp}"
    
    if format_type == 'excel':
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        
        if data:
            # 写入表头
            headers = list(data[0].keys())
            ws.append(headers)
            
            # 写入数据
            for row in data:
                ws.append([row[h] for h in headers])
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = make_response(output.read())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename="{safe_filename}.xlsx"'
        return response
    
    elif format_type == 'csv':
        output = io.StringIO()
        if data:
            writer = csv.DictWriter(output, fieldnames=data[0].keys())
            writer.writeheader()
            writer.writerows(data)
        
        response = make_response(output.getvalue().encode('utf-8-sig'))
        response.headers['Content-Type'] = 'text/csv; charset=utf-8'
        response.headers['Content-Disposition'] = f'attachment; filename="{safe_filename}.csv"'
        return response
    
    elif format_type == 'json':
        response = make_response(json.dumps(data, ensure_ascii=False, indent=2).encode('utf-8'))
        response.headers['Content-Type'] = 'application/json; charset=utf-8'
        response.headers['Content-Disposition'] = f'attachment; filename="{safe_filename}.json"'
        return response
    
    elif format_type == 'txt':
        output = io.StringIO()
        if data:
            headers = list(data[0].keys())
            output.write('\t'.join(headers) + '\n')
            for row in data:
                output.write('\t'.join(str(row[h]) for h in headers) + '\n')
        
        response = make_response(output.getvalue().encode('utf-8'))
        response.headers['Content-Type'] = 'text/plain; charset=utf-8'
        response.headers['Content-Disposition'] = f'attachment; filename="{safe_filename}.txt"'
        return response
    
    return jsonify({'success': False, 'message': '不支持的导出格式'}), 400

@app.route('/teacher/student/<int:student_id>')
def student_detail_page(student_id):
    """学生个人页面"""
    if 'teacher_id' not in session:
        return redirect(url_for('teacher_login'))
    return render_template('student_detail.html', student_id=student_id)

@app.route('/api/teacher/student-stats/<int:student_id>')
def get_student_stats(student_id):
    """获取学生统计数据"""
    if 'teacher_id' not in session:
        return jsonify({'success': False, 'message': '未登录'}), 401
    
    teacher_id = session.get('teacher_id')
    student = db.session.get(Student, student_id)
    
    if not student:
        return jsonify({'success': False, 'message': '学生不存在'}), 404
    
    # 获取该教师布置的所有作业
    homeworks = Homework.query.filter_by(teacher_id=teacher_id).all()
    total_homework = len(homeworks)
    
    # 统计各种状态
    submitted = 0
    approved = 0
    rejected = 0
    reviewing = 0
    error = 0
    
    homework_details = []
    
    for hw in homeworks:
        submission = HomeworkSubmission.query.filter_by(
            student_id=student_id,
            homework_id=hw.id
        ).first()
        
        status = '未提交'
        status_class = 'not-submitted'
        
        if submission:
            submitted += 1
            if submission.ai_review_status == 'approved':
                approved += 1
                status = '已提交-AI审核通过'
                status_class = 'approved'
            elif submission.ai_review_status == 'rejected':
                rejected += 1
                status = '已提交-AI判定异常'
                status_class = 'rejected'
            elif submission.ai_review_status == 'reviewing':
                reviewing += 1
                status = '已提交-AI判定中'
                status_class = 'reviewing'
            elif submission.ai_review_status == 'error':
                error += 1
                status = '已提交-AI审核失败'
                status_class = 'error'
            else:
                status = '已提交'
                status_class = 'submitted'
        
        homework_details.append({
            'homework_id': hw.id,
            'title': hw.title,
            'subject': hw.subject,
            'created_at': hw.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            'status': status,
            'status_class': status_class,
            'submitted': submission is not None
        })
    
    submission_rate = round(submitted / total_homework * 100, 1) if total_homework > 0 else 0
    
    return jsonify({
        'student': {
            'id': student.id,
            'name': student.name,
            'student_id': student.student_id
        },
        'stats': {
            'total_homework': total_homework,
            'submitted': submitted,
            'not_submitted': total_homework - submitted,
            'approved': approved,
            'rejected': rejected,
            'reviewing': reviewing,
            'error': error,
            'submission_rate': submission_rate
        },
        'homework_details': homework_details
    })

# ==================== 管理端路由 ====================
@app.route('/admin')
def admin_index():
    """管理端首页"""
    if 'admin_id' not in session:
        return redirect(url_for('admin_login'))
    return render_template('admin.html')

@app.route('/admin/login')
def admin_login():
    """管理员登录页面"""
    return render_template('admin_login.html')

@app.route('/api/admin/login', methods=['POST'])
def api_admin_login():
    """管理员登录接口"""
    data = request.get_json()
    username = data.get('username')
    password = data.get('password')
    
    if not username or not password:
        return jsonify({'success': False, 'message': '用户名和密码不能为空'}), 400
    
    admin = Admin.query.filter_by(username=username).first()
    if not admin or not check_password_hash(admin.password, password):
        return jsonify({'success': False, 'message': '用户名或密码错误'}), 401
    
    session['admin_id'] = admin.id
    session['admin_username'] = admin.username
    
    return jsonify({'success': True, 'message': '登录成功'}), 200

@app.route('/api/admin/logout', methods=['POST'])
def api_admin_logout():
    """管理员登出接口"""
    session.pop('admin_id', None)
    session.pop('admin_username', None)
    return jsonify({'success': True, 'message': '登出成功'}), 200

@app.route('/api/admin/current-info')
def get_admin_info():
    """获取当前管理员信息"""
    if 'admin_id' not in session:
        return jsonify({'success': False, 'message': '未登录'}), 401

    return jsonify({
        'success': True,
        'username': session.get('admin_username')
    })

# 教师管理
@app.route('/api/admin/teachers')
def get_all_teachers():
    """获取所有教师"""
    if 'admin_id' not in session:
        return jsonify({'success': False, 'message': '未登录'}), 401
    
    teachers = Teacher.query.all()
    teacher_list = []
    for teacher in teachers:
        homework_count = Homework.query.filter_by(teacher_id=teacher.id).count()
        teacher_list.append({
            'id': teacher.id,
            'username': teacher.username,
            'subject': teacher.subject,
            'enable_ai_review': teacher.enable_ai_review,
            'homework_count': homework_count,
            'created_at': teacher.created_at.strftime('%Y-%m-%d %H:%M:%S')
        })
    
    return jsonify(teacher_list)

@app.route('/api/admin/add-teacher', methods=['POST'])
def admin_add_teacher():
    """管理员添加教师"""
    if 'admin_id' not in session:
        return jsonify({'success': False, 'message': '未登录'}), 401
    
    data = request.get_json()
    username = data.get('username')
    password = data.get('password')
    subject = data.get('subject')
    
    if not username or not password or not subject:
        return jsonify({'success': False, 'message': '所有字段都不能为空'}), 400
    
    if len(password) < 6:
        return jsonify({'success': False, 'message': '密码长度至少6位'}), 400
    
    # 检查用户名是否已存在
    existing_teacher = Teacher.query.filter_by(username=username).first()
    if existing_teacher:
        return jsonify({'success': False, 'message': '用户名已存在'}), 400
    
    try:
        hashed_password = generate_password_hash(password)
        teacher = Teacher(username=username, password=hashed_password, subject=subject)
        db.session.add(teacher)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': '教师添加成功',
            'teacher': {
                'id': teacher.id,
                'username': teacher.username,
                'subject': teacher.subject,
                'created_at': teacher.created_at.strftime('%Y-%m-%d %H:%M:%S')
            }
        }), 200
    except Exception as e:
        db.session.rollback()
        print(f"添加教师失败: {str(e)}")
        return jsonify({'success': False, 'message': '添加失败,请重试'}), 500

@app.route('/api/admin/edit-teacher/<int:teacher_id>', methods=['PUT'])
def admin_edit_teacher(teacher_id):
    """管理员编辑教师"""
    if 'admin_id' not in session:
        return jsonify({'success': False, 'message': '未登录'}), 401
    
    teacher = Teacher.query.get(teacher_id)
    if not teacher:
        return jsonify({'success': False, 'message': '教师不存在'}), 404
    
    data = request.get_json()
    username = data.get('username')
    password = data.get('password')
    subject = data.get('subject')
    
    if not username or not subject:
        return jsonify({'success': False, 'message': '用户名和学科不能为空'}), 400
    
    # 检查用户名是否被其他教师使用
    existing_teacher = Teacher.query.filter(Teacher.username == username, Teacher.id != teacher_id).first()
    if existing_teacher:
        return jsonify({'success': False, 'message': '用户名已被其他教师使用'}), 400
    
    try:
        teacher.username = username
        teacher.subject = subject
        if password:  # 如果提供了密码，则更新
            if len(password) < 6:
                return jsonify({'success': False, 'message': '密码长度至少6位'}), 400
            teacher.password = generate_password_hash(password)
        
        db.session.commit()
        return jsonify({'success': True, 'message': '教师信息更新成功'}), 200
    except Exception as e:
        db.session.rollback()
        print(f"编辑教师失败: {str(e)}")
        return jsonify({'success': False, 'message': '更新失败,请重试'}), 500

@app.route('/api/admin/delete-teacher/<int:teacher_id>', methods=['DELETE'])
def admin_delete_teacher(teacher_id):
    """管理员删除教师"""
    if 'admin_id' not in session:
        return jsonify({'success': False, 'message': '未登录'}), 401
    
    teacher = Teacher.query.get(teacher_id)
    if not teacher:
        return jsonify({'success': False, 'message': '教师不存在'}), 404
    
    try:
        # 删除该教师布置的所有作业和相关提交
        homeworks = Homework.query.filter_by(teacher_id=teacher_id).all()
        for hw in homeworks:
            submissions = HomeworkSubmission.query.filter_by(homework_id=hw.id).all()
            for sub in submissions:
                images = HomeworkImage.query.filter_by(submission_id=sub.id).all()
                for img in images:
                    filepath = os.path.join(app.config['UPLOAD_FOLDER'], img.filename)
                    if os.path.exists(filepath):
                        os.remove(filepath)
                    db.session.delete(img)
                db.session.delete(sub)
            db.session.delete(hw)
        
        # 删除教师
        db.session.delete(teacher)
        db.session.commit()
        
        return jsonify({'success': True, 'message': '教师删除成功'}), 200
    except Exception as e:
        db.session.rollback()
        print(f"删除教师失败: {str(e)}")
        return jsonify({'success': False, 'message': '删除失败,请重试'}), 500

# 学生管理
@app.route('/api/admin/students')
def get_all_students_admin():
    """管理员获取所有学生"""
    if 'admin_id' not in session:
        return jsonify({'success': False, 'message': '未登录'}), 401
    
    students = Student.query.all()
    student_list = []
    for student in students:
        submission_count = HomeworkSubmission.query.filter_by(student_id=student.id).count()
        student_list.append({
            'id': student.id,
            'name': student.name,
            'student_id': student.student_id,
            'submission_count': submission_count,
            'created_at': student.created_at.strftime('%Y-%m-%d %H:%M:%S')
        })
    
    return jsonify(student_list)

@app.route('/api/admin/add-student', methods=['POST'])
def admin_add_student():
    """管理员添加学生"""
    if 'admin_id' not in session:
        return jsonify({'success': False, 'message': '未登录'}), 401
    
    data = request.get_json()
    name = data.get('name')
    student_id = data.get('student_id')
    
    if not name or not student_id:
        return jsonify({'success': False, 'message': '姓名和学号不能为空'}), 400
    
    # 检查学号是否已存在
    existing_student = Student.query.filter_by(student_id=student_id).first()
    if existing_student:
        return jsonify({'success': False, 'message': '学号已存在'}), 400
    
    try:
        student = Student(name=name, student_id=student_id)
        db.session.add(student)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': '学生添加成功',
            'student': {
                'id': student.id,
                'name': student.name,
                'student_id': student.student_id,
                'created_at': student.created_at.strftime('%Y-%m-%d %H:%M:%S')
            }
        }), 200
    except Exception as e:
        db.session.rollback()
        print(f"添加学生失败: {str(e)}")
        return jsonify({'success': False, 'message': '添加失败,请重试'}), 500

@app.route('/api/admin/edit-student/<int:student_id>', methods=['PUT'])
def admin_edit_student(student_id):
    """管理员编辑学生"""
    if 'admin_id' not in session:
        return jsonify({'success': False, 'message': '未登录'}), 401
    
    student = Student.query.get(student_id)
    if not student:
        return jsonify({'success': False, 'message': '学生不存在'}), 404
    
    data = request.get_json()
    name = data.get('name')
    new_student_id = data.get('student_id')
    
    if not name or not new_student_id:
        return jsonify({'success': False, 'message': '姓名和学号不能为空'}), 400
    
    # 检查学号是否被其他学生使用
    existing_student = Student.query.filter(Student.student_id == new_student_id, Student.id != student_id).first()
    if existing_student:
        return jsonify({'success': False, 'message': '学号已被其他学生使用'}), 400
    
    try:
        student.name = name
        student.student_id = new_student_id
        db.session.commit()
        return jsonify({'success': True, 'message': '学生信息更新成功'}), 200
    except Exception as e:
        db.session.rollback()
        print(f"编辑学生失败: {str(e)}")
        return jsonify({'success': False, 'message': '更新失败,请重试'}), 500

@app.route('/api/admin/delete-student/<int:student_id>', methods=['DELETE'])
def admin_delete_student(student_id):
    """管理员删除学生"""
    if 'admin_id' not in session:
        return jsonify({'success': False, 'message': '未登录'}), 401
    
    student = Student.query.get(student_id)
    if not student:
        return jsonify({'success': False, 'message': '学生不存在'}), 404
    
    try:
        # 删除该学生的所有作业图片
        submissions = HomeworkSubmission.query.filter_by(student_id=student_id).all()
        for sub in submissions:
            images = HomeworkImage.query.filter_by(submission_id=sub.id).all()
            for img in images:
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], img.filename)
                if os.path.exists(filepath):
                    os.remove(filepath)
                db.session.delete(img)
        
        # 删除该学生的所有作业提交记录
        HomeworkSubmission.query.filter_by(student_id=student_id).delete()
        
        # 删除学生
        db.session.delete(student)
        db.session.commit()
        
        return jsonify({'success': True, 'message': '学生删除成功'}), 200
    except Exception as e:
        db.session.rollback()
        print(f"删除学生失败: {str(e)}")
        return jsonify({'success': False, 'message': '删除失败,请重试'}), 500

@app.route('/api/admin/import-students', methods=['POST'])
def admin_import_students():
    """管理员导入学生Excel文件"""
    if 'admin_id' not in session:
        return jsonify({'success': False, 'message': '未登录'}), 401
    
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '未上传文件'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': '未选择文件'}), 400
    
    # 检查文件类型
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'success': False, 'message': '只支持Excel文件(.xlsx, .xls)'}), 400
    
    try:
        from openpyxl import load_workbook
        import io
        
        # 读取Excel文件
        file_content = file.read()
        workbook = load_workbook(io.BytesIO(file_content))
        sheet = workbook.active
        
        added_count = 0
        skipped_count = 0
        error_rows = []
        
        # 跳过表头，从第2行开始
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or len(row) < 2:
                continue
            
            name = str(row[0]).strip() if row[0] else ''
            student_id = str(row[1]).strip() if row[1] else ''
            
            if not name or not student_id:
                error_rows.append(f"第{row_num}行: 姓名或学号为空")
                continue
            
            # 检查学号是否已存在
            existing_student = Student.query.filter_by(student_id=student_id).first()
            if existing_student:
                skipped_count += 1
                continue
            
            try:
                student = Student(name=name, student_id=student_id)
                db.session.add(student)
                added_count += 1
            except Exception as e:
                error_rows.append(f"第{row_num}行: {str(e)}")
        
        db.session.commit()
        
        result_message = f"成功添加{added_count}个学生"
        if skipped_count > 0:
            result_message += f", 跳过{skipped_count}个重复学号"
        if error_rows:
            result_message += f", {len(error_rows)}个错误"
        
        return jsonify({
            'success': True,
            'message': result_message,
            'added': added_count,
            'skipped': skipped_count,
            'errors': error_rows
        }), 200
        
    except Exception as e:
        db.session.rollback()
        print(f"导入学生失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'导入失败: {str(e)}'}), 500

# 作业管理
@app.route('/api/admin/homeworks')
def get_all_homeworks_admin():
    """管理员获取所有作业"""
    if 'admin_id' not in session:
        return jsonify({'success': False, 'message': '未登录'}), 401
    
    homeworks = Homework.query.order_by(Homework.created_at.desc()).all()
    homework_list = []
    
    for hw in homeworks:
        teacher = hw.teacher
        total_students = Student.query.count()
        submitted_count = HomeworkSubmission.query.filter_by(homework_id=hw.id).count()
        
        homework_list.append({
            'id': hw.id,
            'title': hw.title,
            'subject': hw.subject,
            'teacher_name': teacher.username,
            'teacher_id': teacher.id,
            'created_at': hw.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            'total_students': total_students,
            'submitted_count': submitted_count,
            'max_images': hw.max_images,
            'ai_prompt': hw.ai_prompt
        })
    
    return jsonify(homework_list)

@app.route('/api/admin/delete-homework/<int:homework_id>', methods=['DELETE'])
def admin_delete_homework(homework_id):
    """管理员删除作业"""
    if 'admin_id' not in session:
        return jsonify({'success': False, 'message': '未登录'}), 401
    
    homework = Homework.query.get(homework_id)
    if not homework:
        return jsonify({'success': False, 'message': '作业不存在'}), 404
    
    try:
        # 删除相关图片
        submissions = HomeworkSubmission.query.filter_by(homework_id=homework_id).all()
        for sub in submissions:
            images = HomeworkImage.query.filter_by(submission_id=sub.id).all()
            for img in images:
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], img.filename)
                if os.path.exists(filepath):
                    os.remove(filepath)
                db.session.delete(img)
        
        # 删除相关提交记录
        HomeworkSubmission.query.filter_by(homework_id=homework_id).delete()
        # 删除作业
        db.session.delete(homework)
        db.session.commit()
        
        return jsonify({'success': True, 'message': '作业删除成功'}), 200
    except Exception as e:
        db.session.rollback()
        print(f"删除作业失败: {str(e)}")
        return jsonify({'success': False, 'message': '删除失败,请重试'}), 500

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5009)



